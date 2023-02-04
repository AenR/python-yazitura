#<----------imports---------->
import random #yazı/tura sonucunu bulmak için
import time #sleep komutu için
import os #konsolu temizlemek için
from pyftext import Text
from openpyxl import Workbook,load_workbook #bakiyemizin programı kapatınca sıfırlanmaması için
#<----------/imports---------->

#<----------variables---------->
def temizle():
    os.system('cls' if os.name=='nt' else 'clear')
wb = load_workbook("bakiye.xlsx")
ws = wb.active
sonuc = random.randint(1,2) #1 ve 2 arasında sonuç oluşturma
bakiye= ws["A1"] #bakiyeyi excel dosyasının A1 hücresine eşitleme
wb.close()
#<----------/variables---------->

#<----------splashscreen---------->
hosgeldin = Text(text="AenR", fontsize=10)
hosgeldin.rastgele()
time.sleep(1)
temizle()
#<----------/splashscreen---------->
print("Hoş geldiniz güncel bakiyeniz: {}".format(bakiye.value))
while True: #her oyun sonrası programı yeniden başlatmamak için
    bahis = int(input("Ne kadar bahis yapacaksın?\n0-) Çıkış\n"))
    if(bahis>bakiye.value): #Bakiyenin -'ye düşmemesi için
        print("Yeterince paranız yok.")
        break
    if(bahis==0): #Programdan çıkış için
        print("Hoşça kal")
        break
    bakiye.value -= bahis
    ws["A1"]=bakiye.value
    wb.save("bakiye.xlsx")

    sec = input("Yazı mı, tura mı?\n")
    if(sec=="yazi" or "yazı"):
        yazitura = 1
    if(sec=="tura"):
        yazitura = 2
    if(sonuc==yazitura):
        bahis *= 2
        bakiye.value += bahis
        print("Tutturdun yeni bakiyen: ",bakiye.value)
        ws["A1"]=bakiye.value
        wb.save("bakiye.xlsx")
    else:
        print("Şansına küs yeni bakiyen: ",bakiye.value)